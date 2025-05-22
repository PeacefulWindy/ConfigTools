import os
import sys
import openpyxl
import json
from slpp import slpp as lua
import xml.etree.ElementTree as ET
import time
import shutil
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import colorama
import re

def exportToLua(sheetName, data, outputDir):
    outputFileName = os.path.join(outputDir, f"{sheetName}.lua")
    with open(outputFileName, 'w', encoding='utf-8') as f:
        luaStr = lua.encode(data)
        f.write("return "+luaStr)

def exportToJson(sheetName, data, outputDir):
    outputFileName = os.path.join(outputDir, f"{sheetName}.json")
    with open(outputFileName, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=None,separators=(",", ":"))

def exportDictToXml(parent, data):
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, dict):
                subElement = ET.SubElement(parent, str(key))
                exportDictToXml(subElement, value)
            else:
                subElement = ET.SubElement(parent, str(key))
                subElement.text = str(value)

def exportToXml(sheetName, data, outputDir):
    outputFileName = os.path.join(outputDir, f"{sheetName}.xml")
    root = ET.Element("xml")
    for key, value in data.items():
        exportDictToXml(root, {key: value})
    tree = ET.ElementTree(root)
    tree.write(outputFileName, encoding='utf-8')

formatFunc={
    "lua":exportToLua,
    "json":exportToJson,
    "xml":exportToXml,
}

keyPattern = re.compile(r'^[A-Za-z0-9]+$')

sheets={}

def generateExcel(filePath,outputPath):
    print("--------------------")
    print("open %s"%filePath)

    wb = openpyxl.load_workbook(filePath,data_only=True)

    outputClientRootDir = os.path.join(outputPath, "client")
    outputServerRootDir = os.path.join(outputPath, "server")

    sheetNames=[]
    for sheet in wb.worksheets:
        if not sheet.title.startswith("#"):
            continue
        
        sheetName = sheet.title[1:]
        if sheetName in sheets:
            print("%srepeat sheet:%s\nfile:%s\nanotherFile:%s\n" % (colorama.Fore.RED,sheetName,sheets[sheetName],filePath))
            return
        
        sheets[sheetName]=filePath
        print("sheetName:%s" % sheetName)
        
        idKey = sheet.cell(row=2, column=1).value

        for ext,func in formatFunc.items():
            outputClientDir = os.path.join(outputClientRootDir, ext)
            if not os.path.exists(outputClientDir):
                os.makedirs(outputClientDir)
            
            outputServerDir = os.path.join(outputServerRootDir, ext)
            if not os.path.exists(outputServerDir):
                os.makedirs(outputServerDir)
            
            clientData = {}
            serverData = {}

            ids=[]
            
            for row in range(4, sheet.max_row + 1):
                id = sheet.cell(row=row, column=1).value

                if not id:
                    continue

                if idKey == "int":
                    id=int(id)

                if id in ids:
                    print("%srepeat id:%s\nfile:%s\nsheetName:%s\n" % (colorama.Fore.RED,id,filePath,sheetName))
                    return
                
                ids.append(id)

                clientRowData = {}
                serverRowData = {}

                keys=[]
                for col in range(1, sheet.max_column + 1):
                    colKey = sheet.cell(row=1, column=col).value
                    if not colKey or len(colKey) <= 0:
                        continue

                    if colKey in keys:
                        print("%srepeat key:%s\nfile:%s\nsheetName:%s\nid:%s" % (colorama.Fore.RED,id,filePath,sheetName,str(id)))
                        return

                    keyFirstChar=colKey[0]

                    if keyFirstChar == "#":
                        continue

                    if keyFirstChar == "!" or keyFirstChar == "$":
                        key=colKey[1:]
                    else:
                        key=colKey
                    
                    if not bool(keyPattern.fullmatch(key)):
                        print("%sinvalid key:%s\nfile:%s\nsheet:%s" % (colorama.Fore.RED,key,filePath,sheetName))
                        return

                    colType = sheet.cell(row=2,column=col).value
                    colValue = sheet.cell(row=row, column=col).value

                    if colType == "json":
                        if colValue:
                            try:
                                colValue=json.loads(colValue)
                            except ValueError as e:
                                print("%sinvalid json:%s\nfile:%s\nsheet:%s\nid:%s\nkey:%s" % (colorama.Fore.RED,colValue,filePath,sheetName,id,key))
                                return
                        else:
                            colValue={}
                    elif colType == "int":
                        if not colValue:
                            colValue=0
                        else:
                            colValue=int(colValue)
                    elif colType == "float":
                        if not colValue:
                            colValue=0.0
                        else:
                            colValue=float(colValue)
                    elif colType == "string":
                        if not colValue:
                            colValue=""
                    elif colType == "bool":
                        if not colValue:
                            colValue=False
                        else:
                            colValue=bool(colValue)
                    
                    if keyFirstChar != "!":
                        clientRowData[key]=colValue
                    
                    if keyFirstChar != "$":
                        serverRowData[key]=colValue

                clientData[id] = clientRowData
                serverData[id] = serverRowData
            
            startTime = time.time()
            func(sheetName, clientData,outputClientDir)
            func(sheetName, serverData,outputServerDir)
            endTime = time.time()
            execTime = endTime - startTime
            print("%s%s generate time:%.2fs" % (colorama.Fore.GREEN,ext,execTime))
        
        sheetNames.append(sheetName)

    print("--------------------")
    print("")
    return sheetNames

def move(configData,target):
    moveRoot=configData["move"]
    if not target in moveRoot:
        return
    
    inputPath=os.path.abspath(os.path.join(configData["output"],target))
    
    for format in moveRoot[target]:
        inputTargetPath=os.path.join(inputPath,format)
        
        for it in moveRoot[target][format]:
            outPutPath=os.path.abspath(it)
            if not os.path.exists(outPutPath):
                os.makedirs(outPutPath)
            
            print("%s => %s" % (inputTargetPath,outPutPath))
            shutil.copytree(inputTargetPath,outPutPath,dirs_exist_ok=True)

def main(configData):
    global sheets
    sheets={}
    
    inputPath=os.path.abspath(configData["input"])
    outputPath=os.path.abspath(configData["output"])

    print("inputPath:%s"%inputPath)
    print("outputPath:%s"%outputPath)
    print()

    print("generate excel")
    print()

    sheetDatas=[]
    for root, dirs, files in os.walk(inputPath):
        for file in files:
            if not file.startswith("~") and file.endswith(".xlsx"):
                filePath = os.path.join(root, file)
                ret=generateExcel(filePath,outputPath)
                if not ret:
                    return
                
                sheetDatas.extend(ret)
    
    for ext,func in formatFunc.items():
        outputClientRootDir = os.path.join(outputPath, "client",ext)
        outputServerBaseDir = os.path.join(outputPath, "server",ext)
        data={}
        
        if os.path.exists(outputClientRootDir):
            for it in sheetDatas:
                if os.path.exists(os.path.join(outputClientRootDir,it+"."+ext)):
                    data[it]=True
            func("init",data,outputClientRootDir)
            data.clear()
        
        if os.path.exists(outputServerBaseDir):
            for it in sheetDatas:
                if os.path.exists(os.path.join(outputServerBaseDir,it+"."+ext)):
                    data[it]=True
            func("init",data,outputServerBaseDir)

    if not "move" in configData:
        return
    
    print("move client config")
    move(configData,"client")
    print("move server config")
    move(configData,"server")
    print()

def run(configData):
    startTime = time.time()
    main(configData)
    endTime = time.time()
    execTime = endTime - startTime
    print("%stotal time:%.2fs" % (colorama.Fore.GREEN,execTime))
    print("done!")

configData={}
class FileMonitor(FileSystemEventHandler):
    def __init__(self):
        self.lastModifiedTimes = {}
    
    def on_modified(self, event):
        if not event.is_directory:
            filePath=event.src_path
            if filePath[0] == "~" or not filePath.endswith(".xlsx"):
                return
            
            curTime=os.path.getmtime(filePath)

            if filePath not in self.lastModifiedTimes or self.lastModifiedTimes[filePath] != curTime:
                self.lastModifiedTimes[filePath] = curTime
                os.system('cls' if os.name == 'nt' else 'clear')
                run(configData)

if __name__ == "__main__":
    colorama.init(autoreset=True)

    configPath=os.path.abspath(sys.argv[1])
    if not os.path.exists(configPath):
        print("%snot found config file:%s" % (colorama.Fore.RED,configPath))
        exit(-1)

    with open(configPath, "r", encoding="utf-8") as file:
        configData = json.load(file)
    
    inputPath=os.path.abspath(configData["input"])
    if not os.path.exists(inputPath):
        print("%sinvalid input path:%s" % (colorama.Fore.RED,inputPath))
        exit(-1)
    
    outputPath=os.path.abspath(configData["output"])
    if not os.path.exists(outputPath):
        os.makedirs(outputPath)
    
    if len(sys.argv) > 2 and sys.argv[2] == "-once":
        run(configData)
    else:
        fileMonitorHandle=FileMonitor()
        fileMonitorObserver = Observer()
        fileMonitorObserver.schedule(fileMonitorHandle, inputPath, recursive=False)
        fileMonitorObserver.start()

        print("%sauto generate config service start!" % colorama.Fore.GREEN)

        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            fileMonitorObserver.stop()
            print("auto generate config service stop!")

        fileMonitorObserver.join()